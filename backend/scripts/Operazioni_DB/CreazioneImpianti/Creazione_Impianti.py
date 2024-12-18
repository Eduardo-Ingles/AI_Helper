
import openpyxl
from openpyxl import Workbook
import pandas as pd

from bson import ObjectId
from bson import json_util
import pymongo

import os, sys
project_root = os.getcwd()
sys.path.append(project_root)


from backend.scripts.SharedCode import sharedCode, mongoSearch
from backend.scripts.Operazioni_DB.CreazioneImpianti import crea_ed_inserisci


# ----------------------------------- START COSTANTI -----------------------------------#  

# ----------------------------------- END COSTANTI -----------------------------------# 



def CaricaDaFileImpianto(**kwargs):                                                                              
    print("\n<---caricamento da impianto--->" )

    fileName = kwargs.get("file")
    filepath = kwargs.get("path")
    
    if(fileName):
        if(".xlsx" not in fileName):
            fileName = fileName + ".xlsx"

    if(filepath != ""):    
        if(not (fileName.startswith("\\") or fileName.startswith("/")) and not (filepath.endswith("\\") or filepath.endswith("/"))):
            filepath += "\\" 
        if(project_root.lower().replace("/","\\") not in filepath.lower().replace("/","\\")):
            if(((project_root.endswith("\\") or project_root.endswith("/")) and not (filepath.startswith("\\") or filepath.startswith("/"))) or
               (not (project_root.endswith("\\") or project_root.endswith("/")) and (filepath.startswith("\\") or filepath.startswith("/")))):
                filepath = project_root + filepath
            elif(not (project_root.endswith("\\") or project_root.endswith("/")) and not (filepath.startswith("\\") or filepath.startswith("/"))):
                filepath = project_root + "\\" + filepath

    if(fileName):                                                                          
        workBook = openpyxl.load_workbook(filepath + fileName)#, read_only = True)
        workSheet = workBook.active
        print(workSheet.max_row)
        offset = 2
        ind = 0
        plantDict = None
        subPlantData = []
        for i in range (0, workSheet.max_row):   
            if(i == 0): 
                workSheet["A" + str(i + offset)].value
                if(workSheet["A" + str(i + offset)].value  != None):        
                    plantDict = {
                        "nomeGalleria"  : workSheet["A" + str(i + offset)].value,
                        "plantDescr"    : workSheet["B" + str(i + offset)].value, 
                        "plantTitle"    : workSheet["C" + str(i + offset)].value,  
                        "plantSubTitle" : workSheet["D" + str(i + offset)].value,  
                        "latitudine"    : workSheet["E" + str(i + offset)].value,  
                        "longitudine"   : workSheet["F" + str(i + offset)].value, 
                        "SOC"           : workSheet["G" + str(i + offset)].value   
                        }  
                    
            elif(i > 0 and plantDict["nomeGalleria"] != None):
                if(workSheet["E" + str(i + offset)].value != None and not "Nome Sottoimpianto" in workSheet["B" + str(i + offset)].value):
                    subPlantDict = {
                        "subPlantName" : plantDict["nomeGalleria"] + "-" + workSheet["B" + str(i + offset)].value if not (workSheet["B" + str(i + offset)].value).startswith(plantDict["nomeGalleria"]) else workSheet["B" + str(i + offset)].value,
                        "subPlantDescr" : workSheet["C" + str(i + offset)].value,
                        "icona" : workSheet["D" + str(i + offset)].value,
                        "posizione" : workSheet["E" + str(i + offset)].value,
                        "larghezza" : workSheet["F" + str(i + offset)].value,
                        "altezza" : workSheet["G" + str(i + offset)].value,
                        "pathBG" : workSheet["H" + str(i + offset)].value,
                        
                        "cabine" :  workSheet["I" + str(i + offset)].value,  
                        "dirSx" :  workSheet["J" + str(i + offset)].value, 
                        "dirDx" :  workSheet["K" + str(i + offset)].value, 

                        "alias" : workSheet["L" + str(i + offset)].value                        
                        }

                    subPlantData.append(subPlantDict) if subPlantDict not in subPlantData else next
                    ind += 1  
        workBook.close()
        return {"plantData": plantDict, "subPlantData": subPlantData}



def creaGruppi():
    0

def creaSubPlant():
    0

def mainCall(filePath, fileName, db_Scelto):
    dbAddress = sharedCode.loadSettings("globalSettings","MongoQualityClient")
    if("prod" in db_Scelto.lower() or "p" in db_Scelto.lower()):    
        dbAddress = sharedCode.loadSettings("globalSettings","MongoProductionClient")

    client = pymongo.MongoClient(dbAddress)
    currDB = client["smartscada"] 

    fileData = CaricaDaFileImpianto(path = filePath, file = fileName)
    print(fileData.keys(), fileData["plantData"]["nomeGalleria"]) # plantData subPlantData

    collectionName = "plant"
    nomeGalleria = fileData["plantData"]["nomeGalleria"]
    plantResult = mongoSearch.readCollectionData(currDB, collectionName, name = nomeGalleria+"_")


    if(plantResult):
        print(f"La galleria {nomeGalleria} esiste già! -> {plantResult.get("name")} - {plantResult.get("_id")}")
        return False
    else: 
        newSubPlants = []
        print(nomeGalleria)

        uid = crea_ed_inserisci.newDocument_Plant(fileData["plantData"], currDB)

        if(not mongoSearch.readCollectionData(currDB, collectionName, id = uid)):

            print(fileData["subPlantData"][0].keys())   
            #creaSubPlants.creaSubPlants(fileData["subPlantData"], galleria = nomeGalleria)  
            newGroupHolder = []
            existingGroupsHolder = []

            newSubPlantHolder = []
            existingPlantHolder = []
            """Creazione GRUPPI"""
            for subPlants in fileData["subPlantData"]:
                nomeGruppo = nomeGalleria.replace(" ","-") + "-" + subPlants["subPlantName"] if(not subPlants["subPlantName"].startswith(nomeGalleria)) else subPlants["subPlantName"]

                searchGroup = mongoSearch.readCollectionData(currDB, "group", name = nomeGruppo)
                if(searchGroup):
                    print(f"Gruppo {nomeGruppo} già esistente! ->\t creato nuovo gruppo:\t {nomeGruppo + "_NEW"}")
                    nomeGruppo += "_NEW"
                    existingGroupsHolder.append(searchGroup) if searchGroup not in existingGroupsHolder else next
                groupDoc = crea_ed_inserisci.newDocument_group(name = nomeGruppo, descrizione = nomeGruppo)
                newGroupHolder.append(groupDoc) if groupDoc not in newGroupHolder else next
                #print(json_util.dumps(groupDoc, indent = 4))


                nomeSubPlant = nomeGalleria.replace(" ","-") + "-" + subPlants["subPlantName"] if(not subPlants["subPlantName"].startswith(nomeGalleria)) else subPlants["subPlantName"]

                searchSp = mongoSearch.readCollectionData(currDB, "subPlant", name = nomeSubPlant)
                if(searchSp):
                    print(f"subPlant {nomeGruppo} già esistente! ->\t creato nuovo subPlant:\t {nomeSubPlant + "_NEW"}")                
                    existingPlantHolder.append(searchSp) if searchSp not in existingPlantHolder else next
                    nomeSubPlant += "_NEW"
                #doc = crea_ed_inserisci.newDocument_subPlant(name = nomeSubPlant, 
                #                                              descrizione = subPlants["subPlantDescr"],
                #                                              parent = "",
                #                                              groupid = groupDoc["_id"],
                #                                              icon = "", 
                #                                              position = "", 
                #                                              order = "", 
                #                                              larghezza = "", 
                #                                              altezza = "", 
                #                                              bg = "", 
                #                                              root = "",
                #                                              sinottico = ""
                #                                              )

            print(len(newGroupHolder), len(existingGroupsHolder))

        #   "name": kwargs.get("name").upper(),
        #   "description": kwargs.get("descrizione"),
        #   parentId": kwargs.get("parent"),
        #   group" : DBRef("group", ObjectId(kwargs.get("groupid"))),
        #   "icon": kwargs.get("icon"),

        #    "devicePositions": kwargs.get("position"),
        #    "order": kwargs.get("order"),
        #    "width": kwargs.get("larghezza"),#1920,
        #    "height": kwargs.get("altezza"),#1200,
        #    "backgroundImg": kwargs.get("bg"),#"parcodellareggia/background/bg_qv1.svg",
        #    "root": kwargs.get("root"),
        #    "main": kwargs.get("sinottico"),


        

        #choice = input("Utilizzare nuovi gruppi oppure già esistenti? (old/new):\n")


        #for subPlat in fileData["subPlantData"]:
            #print(subPlat.get("subPlantName"), subPlat.get("subPlantDescr"))            

            #subPlantName = subPlat["subPlantName"] if (subPlat["subPlantName"].upper().startswith(nomeGalleria.upper())) else nomeGalleria + "-" + subPlat["subPlantName"]



            #print(tempSP)t



if __name__ == '__main__':
    
    filePath = "\\scripts\\TestCode\\miscTestCode"
    fileName = "creazione card galleria.xlsx"


    mainCall(filePath, fileName, "q")
