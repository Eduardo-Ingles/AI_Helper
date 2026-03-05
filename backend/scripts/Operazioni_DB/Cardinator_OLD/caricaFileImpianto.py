#caricaFileImpianto

import openpyxl
from openpyxl import Workbook
from datetime import datetime
from bson import ObjectId
from bson.dbref import DBRef
import os
import sys


project_root = os.getcwd()
sys.path.append(project_root)

arr_colonneIn = [  "A"           #  0: Galleria
                 , "B"           #  1: Nome Sottoimpianto
                 , "C"           #  2: Descrizione
                 , "D"           #  3: Icona
                 , "E"           #  4: posizione
                 , "F"           #  5: larghezza / Width
                 , "G"           #  6: altezza / Height
                 , "H"           #  7: percorso BG / background
                 , "I"           #  8: Alias-Anagrafica/Mappa
                ]
                # , "J"           #  9: Gruppo                
                # , "K"           # 10: show
                # , "L"           # 11: root
				# , "M"           # 12: sinottico
                #]

class DatiImpianto:
    def __init__(self):  
        self.nomeGalleria   = None    
        self.plantDescr     = None     
        self.plantTitle     = None    
        self.plantSubTitle  = None           
        self.latitudine     = None 
        self.longitudine    = None 
        self.SOC            = None 
        self.subPlants      = None 

        self.subPlantName   = None 
        self.subPlantDescr  = None 
        self.icona          = None
        self.posizione      = None
        self.larghezza      = None
        self.altezza        = None
        self.pathBG         = None
        self.alias          = None
        
        self.cabine         = None
        self.dirSx          = None
        self.dirDx          = None
        #self.sinottico      = None

def CaricaDaFileImpianto(file):                                                                                   # <<<<< used function here <<<<
    print("\n<---caricamento da impianto--->" )
    arrDatiPlant = []
    if(not "N/D" in file):                                                                              
        workBook = openpyxl.load_workbook(file)#, read_only = True)
        workSheet = workBook.active
        offset = 2
        ind = 0
        for i in range (0, workSheet.max_row):   
            if(i == 0): 
                if(workSheet["A" + str(i + offset)].value  != None): 
                    arrDatiPlant.append(DatiImpianto())               
                    arrDatiPlant[ind].nomeGalleria  = workSheet["A" + str(i + offset)].value  
                    arrDatiPlant[ind].plantDescr    = workSheet["B" + str(i + offset)].value  
                    arrDatiPlant[ind].plantTitle    = workSheet["C" + str(i + offset)].value  
                    arrDatiPlant[ind].plantSubTitle = workSheet["D" + str(i + offset)].value  
                    arrDatiPlant[ind].latitudine    = workSheet["E" + str(i + offset)].value  
                    arrDatiPlant[ind].longitudine   = workSheet["F" + str(i + offset)].value  
                    arrDatiPlant[ind].SOC           = workSheet["G" + str(i + offset)].value  
                    arrDatiPlant[ind].wip           = workSheet["M" + str(i + offset)].value  
                    #arrDatiPlant[ind].subPlants     = workSheet["H" + str(i + offset)].value 
            elif(i > 0 and arrDatiPlant[0].nomeGalleria != None):
                if(workSheet["E" + str(i + offset)].value != None and not "Nome Sottoimpianto" in workSheet["B" + str(i + offset)].value):
                    arrDatiPlant.append(DatiImpianto()) 
                    ind += 1
                    arrDatiPlant[ind].subPlantName  = workSheet["B" + str(i + offset)].value  
                    arrDatiPlant[ind].subPlantDescr = workSheet["C" + str(i + offset)].value  
                    arrDatiPlant[ind].icona         = workSheet["D" + str(i + offset)].value  
                    arrDatiPlant[ind].posizione     = workSheet["E" + str(i + offset)].value  
                    arrDatiPlant[ind].larghezza     = workSheet["F" + str(i + offset)].value  
                    arrDatiPlant[ind].altezza       = workSheet["G" + str(i + offset)].value  
                    arrDatiPlant[ind].pathBG        = workSheet["H" + str(i + offset)].value  
                    
                    arrDatiPlant[ind].cabine        = workSheet["I" + str(i + offset)].value  
                    arrDatiPlant[ind].dirSx         = workSheet["J" + str(i + offset)].value 
                    arrDatiPlant[ind].dirDx         = workSheet["K" + str(i + offset)].value 

                    arrDatiPlant[ind].alias         = workSheet["L" + str(i + offset)].value  
                    #arrDatiPlant[ind].sinottico     = workSheet["M" + str(i + offset)].value 
            #print(f"{ind}|{i}: {arrDatiPlant[ind].subPlantName}")
        workBook.close()
    #printaFunc(arrDatiPlant)
    return arrDatiPlant


def SalvaDaFileImpianto(file):                                                                                   # <<<<< used function here <<<<
    print("\n<---caricamento da impianto--->" )
    arrDatiPlant = []
    if(not "N/D" in file):      
        workBook = Workbook()
        workSheet = workBook.active
        workSheet.title = "New Title"
        offset = 2
        ind = 0
        for i in range (0, workSheet.max_row):   
            if(i == 0): 
                if(workSheet["A" + str(i + offset)].value  != None): 
                    arrDatiPlant.append(DatiImpianto())               
                    arrDatiPlant[ind].nomeGalleria  = workSheet["A" + str(i + offset)].value  
                    arrDatiPlant[ind].plantDescr    = workSheet["B" + str(i + offset)].value  
                    arrDatiPlant[ind].plantTitle    = workSheet["C" + str(i + offset)].value  
                    arrDatiPlant[ind].plantSubTitle = str(workSheet["D" + str(i + offset)].value)  
                    arrDatiPlant[ind].latitudine    = str(workSheet["E" + str(i + offset)].value)  
                    arrDatiPlant[ind].longitudine   = str(workSheet["F" + str(i + offset)].value)  
                    arrDatiPlant[ind].SOC           = str(workSheet["G" + str(i + offset)].value)  
                    arrDatiPlant[ind].wip           = str(workSheet["M" + str(i + offset)].value)  
                    arrDatiPlant[ind].subPlants     = workSheet["H" + str(i + offset)].value 
            elif(i > 0 and arrDatiPlant[0].nomeGalleria != None):
                if(workSheet["E" + str(i + offset)].value != None and not "Nome Sottoimpianto" in workSheet["B" + str(i + offset)].value):
                    arrDatiPlant.append(DatiImpianto()) 
                    ind += 1
                    arrDatiPlant[ind].subPlantName  = str(workSheet["B" + str(i + offset)].value)  
                    arrDatiPlant[ind].subPlantDescr = str(workSheet["C" + str(i + offset)].value)  
                    arrDatiPlant[ind].icona         = str(workSheet["D" + str(i + offset)].value)  
                    arrDatiPlant[ind].posizione     = str(workSheet["E" + str(i + offset)].value)  
                    arrDatiPlant[ind].larghezza     = str(workSheet["F" + str(i + offset)].value)  
                    arrDatiPlant[ind].altezza       = str(workSheet["G" + str(i + offset)].value)  
                    arrDatiPlant[ind].pathBG        = str(workSheet["H" + str(i + offset)].value)  
                    arrDatiPlant[ind].alias         = str(workSheet["I" + str(i + offset)].value)  
                    arrDatiPlant[ind].gruppo        = str(workSheet["J" + str(i + offset)].value)  
                    
        workBook.save('balances.xlsx')
        workBook.close()
    return arrDatiPlant


def printaFunc(arrDatiPlant, **kwargs):    
    for i in range(len(arrDatiPlant)):
        if(i == 0):             
            msg = (  arrDatiPlant[i].nomeGalleria, " "  
                    ,arrDatiPlant[i].plantDescr, " "    
                    ,arrDatiPlant[i].plantTitle, " "    
                    ,arrDatiPlant[i].plantSubTitle, " " 
                    ,arrDatiPlant[i].latitudine, " "    
                    ,arrDatiPlant[i].longitudine, " "   
                    ,arrDatiPlant[i].SOC, " "      
                    ,arrDatiPlant[i].wip, " "       
                    #,arrDatiPlant[i].subPlants, 
                    "\n" )    
            yield msg if "yieldFlag" in kwargs and kwargs.get("yieldFlag") else print(msg)
        elif(i > 0):
            msg = (  arrDatiPlant[i].subPlantName, " " 
                    ,arrDatiPlant[i].subPlantDescr, " " 
                    ,arrDatiPlant[i].icona, " "         
                    ,arrDatiPlant[i].posizione, " "     
                    ,arrDatiPlant[i].larghezza, " "     
                    ,arrDatiPlant[i].altezza, " "       
                    ,arrDatiPlant[i].pathBG, " "        
                    ,arrDatiPlant[i].alias, " ")         
                    #,arrDatiPlant[i].gruppo, " "   
                    #,arrDatiPlant[i].show, " "   
                    #,arrDatiPlant[i].root, " "   
                    #,arrDatiPlant[i].sinottico
                    #)
            yield msg if "yieldFlag" in kwargs and kwargs.get("yieldFlag") else print(msg)
            
if __name__ == '__main__':
    print(project_root)
    arrDatiPlant = CaricaDaFileImpianto(project_root + "\\uploads\\impianto_VALSASSINA.xlsx")  
    printaFunc(arrDatiPlant)